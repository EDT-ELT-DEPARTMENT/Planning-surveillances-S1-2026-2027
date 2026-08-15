import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm
import os
import re

TITRE_PLATEFORME = "Plateforme de gestion des EDTs-S2-2026-Département d'Électrotechnique-Faculté de génie électrique-UDL-SBA"

st.set_page_config(page_title=TITRE_PLATEFORME, page_icon="📋", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
.main-header { font-size: 1.8rem; font-weight: bold; color: #1f4e79; text-align: center; padding: 1rem; background: linear-gradient(90deg, #e3f2fd 0%, #bbdefb 100%); border-radius: 10px; margin-bottom: 1.5rem; }
.sub-header { font-size: 1.4rem; font-weight: bold; color: #1565c0; margin-top: 1rem; margin-bottom: 0.5rem; border-bottom: 2px solid #1565c0; padding-bottom: 0.3rem; }
.info-box { background-color: #e3f2fd; padding: 1rem; border-radius: 8px; border-left: 4px solid #1565c0; margin: 0.5rem 0; text-align: center; }
.stTabs [data-baseweb="tab-list"] { gap: 8px; justify-content: center; }
.stTabs [data-baseweb="tab"] { background-color: #f5f5f5; border-radius: 8px 8px 0 0; padding: 10px 20px; font-weight: 600; text-align: center; }
.stTabs [aria-selected="true"] { background-color: #1565c0 !important; color: white !important; }
p, div, span, th, td { text-align: center !important; }
</style>
""", unsafe_allow_html=True)

SALLES = [f"S{i:02d}" for i in range(1, 18)]
AMPHIS = [f"A{i:02d}" for i in range(1, 13)]
CRENEAUX_DEFAUT = ["08h30 - 10h30", "11h00 - 13h00", "13h30 - 15h30"]
JOURS_FR = {"Monday": "Lundi", "Tuesday": "Mardi", "Wednesday": "Mercredi", "Thursday": "Jeudi", "Friday": "Vendredi", "Saturday": "Samedi", "Sunday": "Dimanche"}
FICHIER_SOURCE = "DATA-ENS-2026-2027_surveillances.xlsx"

# Paramètres SMTP par défaut (hardcodés)
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
SMTP_USER = "chef.department.elt.fge@gmail.com"
SMTP_PASSWORD = "gkzs pdza yodb icvd"

# Fichier de persistance des EDTs
FICHIER_PERSISTANCE = "edt_persistence.pkl"

def nettoyer_nom_feuille(nom):
    """Nettoie le nom pour respecter les contraintes Excel (max 31 caractères et pas de caractères spéciaux interdits)."""
    nom_propre = re.sub(r'[\\/\?\*\[\]:]', '', str(nom))
    return nom_propre[:31]

import pickle

def sauvegarder_persistence():
    """Sauvegarde les données critiques sur disque pour persistance après déconnexion."""
    try:
        data = {
            'planning_df': st.session_state.get('planning_df'),
            'surveillance_df': st.session_state.get('surveillance_df'),
            'historique_edt': st.session_state.get('historique_edt'),
            'enseignants_df': st.session_state.get('enseignants_df'),
            'examens_df': st.session_state.get('examens_df'),
            'promotions_list': st.session_state.get('promotions_list'),
            'permanents_list': st.session_state.get('permanents_list'),
            'vacataires_list': st.session_state.get('vacataires_list'),
            'all_enseignants_list': st.session_state.get('all_enseignants_list'),
            'data_loaded': st.session_state.get('data_loaded'),
            'promo_selected': st.session_state.get('promo_selected'),
            'creneaux_actifs': st.session_state.get('creneaux_actifs'),
            'lieux_par_promo': st.session_state.get('lieux_par_promo'),
            'groupes_par_promo': st.session_state.get('groupes_par_promo'),
            'fractionnement_actif': st.session_state.get('fractionnement_actif'),
            'jours_feries': st.session_state.get('jours_feries'),
            'jours_feries_list': st.session_state.get('jours_feries_list'),
            'exclus_manuels': st.session_state.get('exclus_manuels'),
            'nb_surv_permanent': st.session_state.get('nb_surv_permanent'),
            'nb_surv_vacataire': st.session_state.get('nb_surv_vacataire'),
            'nb_surv_autre': st.session_state.get('nb_surv_autre'),
            'nb_surv_par_salle': st.session_state.get('nb_surv_par_salle'),
            'nb_surv_par_amphi': st.session_state.get('nb_surv_par_amphi'),
            'date_debut_val': st.session_state.get('date_debut_val'),
            'date_fin_val': st.session_state.get('date_fin_val'),
            'nb_par_jour': st.session_state.get('nb_par_jour'),
        }
        with open(FICHIER_PERSISTANCE, 'wb') as f:
            pickle.dump(data, f)
    except Exception as e:
        pass

def charger_persistence():
    """Recharge les données sauvegardées depuis le disque."""
    if os.path.exists(FICHIER_PERSISTANCE):
        try:
            with open(FICHIER_PERSISTANCE, 'rb') as f:
                data = pickle.load(f)
            for key, value in data.items():
                if key not in st.session_state or st.session_state.get(key) is None:
                    st.session_state[key] = value
            return True
        except Exception as e:
            return False
    return False

def init_session_state():
    defaults = {
        'enseignants_df': None, 'examens_df': None, 'planning_df': None,
        'surveillance_df': None, 'nb_surv_permanent': 3, 'nb_surv_vacataire': 2,
        'nb_surv_autre': 1, 'nb_surv_par_salle': 2, 'nb_surv_par_amphi': 3,
        'exclus_manuels': [], 'date_debut_val': date(2026, 11, 1), 'date_fin_val': date(2026, 11, 15),
        'nb_par_jour': 2, 'jours_feries': [], 'promo_selected': None, 'data_loaded': False,
        'promotions_list': [], 'permanents_list': [], 'vacataires_list': [], 'all_enseignants_list': [],
        'ordre_matieres': {}, 'lieux_par_promo': {}, 'horaires_par_matiere': {}, 'jours_par_matiere': {},
        'fractionnement_actif': {}, 'groupes_par_promo': {}, 'historique_edt': {}, 'creneaux_actifs': CRENEAUX_DEFAUT,
        'round_robin_pointer': 0, 'disponibilites_enseignants': {}
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

def normaliser_qualite(val):
    val = str(val).strip().lower()
    if 'vacataire' in val or 'charg' in val or 'contractuel' in val or 'doctorant' in val:
        return 'Vacataire'
    mapping = {
        'permanent': 'Permanent', 'vacataire': 'Vacataire', 'contractuel': 'Contractuel', 'autre': 'Autre',
        'professeur': 'Permanent', 'maitre de conferences': 'Permanent', 'mc': 'Permanent', 'prof': 'Permanent', 'mca': 'Permanent', 'mcb': 'Permanent'
    }
    for k, v in mapping.items():
        if k in val:
            return v
    return 'Permanent'

def est_cours(enseignement_str):
    val = str(enseignement_str).strip()
    return bool(re.match(r'^[Cc][Oo][Uu][Rr][Ss]', val))

def extraire_nom_cours(enseignement_str):
    val = str(enseignement_str).strip()
    nom = re.sub(r'^[Cc][Oo][Uu][Rr][Ss][ \-_:]+', '', val).strip()
    return nom if nom else val

def corriger_fautes_enseignants(nom):
    if not isinstance(nom, str):
        return nom
    nom_clean = nom.strip()
    if nom_clean == "Belhadj":
        return "Belabed"
    elif nom_clean == "Zeghdoudi":
        return "ZEGHOUDI"
    elif nom_clean == "Babali":
        return "Bahlil"
    return nom_clean

def charger_fichier_source_auto():
    try:
        paths_to_try = [FICHIER_SOURCE, os.path.join(os.getcwd(), FICHIER_SOURCE),
            os.path.join(os.path.dirname(os.path.abspath(__file__)), FICHIER_SOURCE),
            os.path.join("/mnt/agents/upload/", FICHIER_SOURCE),
            os.path.join("/mount/src/planning-surveillances-s1-2026-2027/", FICHIER_SOURCE)]
        file_path = None
        for p in paths_to_try:
            if os.path.exists(p):
                file_path = p
                break
        if file_path is None:
            return None, f"Fichier {FICHIER_SOURCE} non trouvé"
        xls = pd.ExcelFile(file_path)
        sheet_names = xls.sheet_names
        
        ens_sheet = None
        for preferred in ['matières', 'EDTCE']:
            if preferred in sheet_names:
                ens_sheet = preferred
                break
                
        if not ens_sheet:
            for sheet in sheet_names:
                df_test = pd.read_excel(file_path, sheet_name=sheet, nrows=5)
                cols_lower = [str(c).lower().strip() for c in df_test.columns]
                has_qualite = any('qualite' in c or 'quality' in c or 'statut' in c or 'grade' in c for c in cols_lower)
                has_enseignements = any('enseignement' in c or 'cours' in c or 'matiere' in c or 'module' in c for c in cols_lower)
                has_nom = any('nom' in c or 'name' in c or 'enseignant' in c for c in cols_lower)
                if has_qualite and has_enseignements and has_nom:
                    ens_sheet = sheet
                    break
        if ens_sheet is None and len(sheet_names) > 0:
            ens_sheet = sheet_names[0]
            
        df_ens = pd.read_excel(file_path, sheet_name=ens_sheet)
        df_ens.columns = [str(col).strip() for col in df_ens.columns]
        cols_orig = list(df_ens.columns)
        cols_lower = [c.lower().strip().replace(' ', '_').replace('-', '_') for c in cols_orig]
        col_map = {}
        for i, c in enumerate(cols_lower):
            if any(x in c for x in ['nom', 'name', 'enseignant', 'prenom_nom', 'nom_prenom', 'professeur']):
                col_map['nom'] = cols_orig[i]
            elif any(x in c for x in ['qualite', 'quality', 'type', 'statut', 'grade', 'categorie', 'situation']):
                col_map['qualite'] = cols_orig[i]
            elif any(x in c for x in ['enseignement', 'cours', 'matiere', 'module', 'discipline', 'ue', 'matieres', 'enseignements']):
                col_map['enseignements'] = cols_orig[i]
            elif any(x in c for x in ['promotion', 'niveau', 'annee', 'class', 'promo', 'niveaux']):
                col_map['promotion'] = cols_orig[i]
        rename_map = {}
        if 'nom' in col_map: rename_map[col_map['nom']] = 'nom'
        if 'qualite' in col_map: rename_map[col_map['qualite']] = 'qualite'
        if 'enseignements' in col_map: rename_map[col_map['enseignements']] = 'Enseignements'
        if 'promotion' in col_map: rename_map[col_map['promotion']] = 'Promotion'
        df_ens = df_ens.rename(columns=rename_map)
        if 'nom' not in df_ens.columns:
            for col in df_ens.columns:
                if df_ens[col].dtype == 'object':
                    sample = df_ens[col].dropna().astype(str)
                    if len(sample) > 0 and sample.str.len().mean() > 3:
                        df_ens['nom'] = df_ens[col]
                        break
        for col in ['nom', 'qualite', 'Enseignements', 'Promotion']:
            if col not in df_ens.columns:
                df_ens[col] = ''
        df_ens = df_ens[df_ens['nom'].notna() & (df_ens['nom'].astype(str).str.strip() != '')].copy()
        
        df_ens['nom'] = df_ens['nom'].apply(lambda x: corriger_fautes_enseignants(str(x)))
        df_ens['qualite'] = df_ens['qualite'].apply(normaliser_qualite)
        
        examens_data = []
        for _, row in df_ens.iterrows():
            raw_ens = str(row.get('Enseignements', ''))
            items = re.split(r'[,;/]+', raw_ens)
            for item in items:
                item = item.strip()
                if item and est_cours(item):
                    nom_cours = extraire_nom_cours(item)
                    code_cours = f"CODE-{abs(hash(nom_cours)) % 9000 + 1000}"
                    examens_data.append({
                        'Enseignements': nom_cours, 
                        'Code': code_cours,
                        'Enseignants': str(row.get('nom', '')).strip(), 
                        'qualite_ens': row.get('qualite', 'Permanent'),
                        'Horaire': CRENEAUX_DEFAUT[0], 'Jours': None, 'Lieu': None,
                        'Promotion': str(row.get('Promotion', '')).strip(),
                        'Groupe': 'Global',
                        'ordre': 999
                    })
        df_exam = pd.DataFrame(examens_data)
        colonnes_attendues = ['Enseignements', 'Code', 'Enseignants', 'Horaire', 'Jours', 'Lieu', 'Promotion', 'Groupe']
        for col in colonnes_attendues:
            if col not in df_exam.columns:
                df_exam[col] = ''
                
        df_exam = df_exam.drop_duplicates(subset=['Enseignements', 'Promotion', 'Groupe', 'Enseignants']).copy()
        df_exam = df_exam.sort_values('Enseignants').drop_duplicates(subset=['Enseignements', 'Promotion', 'Groupe'], keep='first').copy()
        promotions = sorted(df_exam['Promotion'].dropna().astype(str).str.strip().unique().tolist()) if not df_exam.empty else []
        promotions = [p for p in promotions if p != '']
        
        df_ens['is_perm'] = df_ens['qualite'].apply(lambda x: 0 if x == 'Permanent' else 1)
        df_ens = df_ens.sort_values(['is_perm', 'nom']).drop(columns=['is_perm'])
        
        permanents = df_ens[df_ens['qualite'] == 'Permanent']['nom'].dropna().unique().tolist()
        vacataires = df_ens[df_ens['qualite'] == 'Vacataire']['nom'].dropna().unique().tolist()
        all_ens = df_ens['nom'].dropna().unique().tolist()
        return {'enseignants': df_ens, 'examens': df_exam, 'promotions': promotions,
                'permanents': permanents, 'vacataires': vacataires, 'all_enseignants': all_ens, 'sheet_used': ens_sheet}, None
    except Exception as e:
        return None, str(e)

def est_jour_travaille(date_obj, jours_feries):
    if isinstance(date_obj, datetime):
        date_obj = date_obj.date()
    elif isinstance(date_obj, str):
        try:
            date_obj = datetime.strptime(date_obj, "%Y-%m-%d").date()
        except:
            return True
    if not isinstance(date_obj, date):
        return True
    jour_semaine = date_obj.strftime("%A")
    jour_fr = JOURS_FR.get(jour_semaine, jour_semaine)
    if jour_fr in ["Vendredi", "Samedi"]:
        return False
    for jf in jours_feries:
        if isinstance(jf, str):
            try:
                jf = datetime.strptime(jf, "%d/%m/%Y").date()
            except:
                continue
        if isinstance(jf, datetime):
            jf = jf.date()
        if date_obj == jf:
            return False
    return True

def generer_planning_promo(examens_df, promotion, date_debut, date_fin, nb_par_jour, jours_feries, creneaux, lieux, ordre_matieres=None, horaires_matiere=None, jours_matiere=None, groupes_actifs=None):
    if examens_df is None or examens_df.empty:
        return None
    
    df_working = examens_df.copy()
    promo_df = df_working[df_working['Promotion'].astype(str).str.strip() == str(promotion).strip()].copy()
    if promo_df.empty:
        return examens_df
        
    if groupes_actifs and len(groupes_actifs) > 0 and 'Groupe' in promo_df.columns:
        promo_df = promo_df[promo_df['Groupe'].isin(groupes_actifs + ['Global'])].copy()
    
    if ordre_matieres and promotion in ordre_matieres:
        ordre_map = {m: i for i, m in enumerate(ordre_matieres[promotion])}
        promo_df['ordre'] = promo_df['Enseignements'].map(ordre_map).fillna(999).astype(int)
        promo_df = promo_df.sort_values('ordre')
    else:
        promo_df = promo_df.sort_values('Enseignements')
        
    nb_lieux = len(lieux)
    if nb_lieux == 0:
        st.error("Veuillez sélectionner au moins un lieu.")
        return None
    creneaux_dispo = creneaux if creneaux else CRENEAUX_DEFAUT
    if len(creneaux_dispo) == 0:
        st.error("Veuillez sélectionner au moins un créneau.")
        return None
        
    groupes_matieres = promo_df.groupby(['Code', 'Enseignements'])

    creneaux_occupes = set()
    examens_par_jour_count = {}
    date_courante = date_debut
    lieu_idx = 0

    for (code_m, matiere_nom), group_rows in groupes_matieres:
        date_pref = jours_matiere.get(promotion, {}).get(matiere_nom) if jours_matiere else None
        if date_pref:
            if isinstance(date_pref, datetime):
                d_ex = date_pref.date()
            elif isinstance(date_pref, date):
                d_ex = date_pref
            else:
                try:
                    d_ex = datetime.strptime(str(date_pref), "%Y-%m-%d").date()
                except:
                    d_ex = date_debut
        else:
            d_ex = None
            
        creneau_pref = horaires_matiere.get(promotion, {}).get(matiere_nom) if horaires_matiere else None
        
        found_slot = False
        date_examen = d_ex if d_ex else date_courante
        
        while not found_slot:
            while not est_jour_travaille(date_examen, jours_feries) or date_examen > date_fin:
                date_examen += timedelta(days=1)
                
            exams_ce_jour = examens_par_jour_count.get(date_examen, 0)
            if exams_ce_jour < nb_par_jour or d_ex:
                creneaux_a_tester = [creneau_pref] if (creneau_pref and creneau_pref in creneaux_dispo) else creneaux_dispo
                for c in creneaux_a_tester:
                    if (date_examen, c) not in creneaux_occupes:
                        creneau = c
                        found_slot = True
                        break
            
            if not found_slot:
                date_examen += timedelta(days=1)

        creneaux_occupes.add((date_examen, creneau))
        examens_par_jour_count[date_examen] = examens_par_jour_count.get(date_examen, 0) + 1

        for idx_row in group_rows.index:
            lieu = lieux[lieu_idx % nb_lieux]
            lieu_idx += 1
            
            mask_m = (df_working['Promotion'].astype(str).str.strip() == str(promotion).strip()) & (df_working['Enseignements'] == matiere_nom)
            if 'Code' in df_working.columns:
                mask_m = mask_m & (df_working['Code'] == code_m)
            if 'Groupe' in df_working.columns and promo_df.loc[idx_row, 'Groupe'] in df_working['Groupe'].values:
                mask_m = mask_m & (df_working['Groupe'] == promo_df.loc[idx_row, 'Groupe'])

            df_working.loc[mask_m, 'date'] = date_examen
            df_working.loc[mask_m, 'Horaire'] = creneau
            df_working.loc[mask_m, 'Jours'] = JOURS_FR.get(date_examen.strftime("%A"), date_examen.strftime("%A"))
            df_working.loc[mask_m, 'Lieu'] = lieu

        if not d_ex:
            date_courante = date_examen
            
    return df_working

def attribuer_surveillants(planning_df, enseignants_df):
    if planning_df is None or enseignants_df is None:
        return None, enseignants_df
        
    surveillants = enseignants_df.copy()
    if 'surveillance_attribuee' not in surveillants.columns:
        surveillants['surveillance_attribuee'] = 0
        
    exclus = st.session_state.get('exclus_manuels', [])
    attributions = []
    
    if 'disponibilites_enseignants' not in st.session_state or not isinstance(st.session_state.disponibilites_enseignants, dict):
        st.session_state.disponibilites_enseignants = {}
    disponibilites_enseignants = st.session_state.disponibilites_enseignants
    
    for ens_nom in surveillants['nom'].unique():
        if ens_nom not in disponibilites_enseignants:
            disponibilites_enseignants[ens_nom] = set()

    liste_tous_ens = surveillants[~surveillants['nom'].isin(exclus)]['nom'].tolist()
    if not liste_tous_ens:
        liste_tous_ens = surveillants['nom'].tolist()
    
    if 'round_robin_pointer' not in st.session_state:
        st.session_state.round_robin_pointer = 0

    creneaux_defaut_locaux = ['08h30 - 10h30']
    try:
        creneaux_actifs = st.session_state.get('creneaux_actifs', CRENEAUX_DEFAUT)
    except NameError:
        creneaux_actifs = st.session_state.get('creneaux_actifs', creneaux_defaut_locaux)

    charges_affectees_creneau = set()

    def trouver_surveillant_round_robin(qualite_souhaitee=None):
        n_total = len(liste_tous_ens)
        if n_total == 0:
            return None, None
            
        for i in range(n_total):
            ptr = (st.session_state.round_robin_pointer + i) % n_total
            s_nom = liste_tous_ens[ptr]
            
            if s_nom in exclus:
                continue
            
            row_s = surveillants[surveillants['nom'] == s_nom]
            if row_s.empty:
                continue
            q_val = row_s.iloc[0]['qualite']
            
            if qualite_souhaitee and q_val != qualite_souhaitee:
                continue
            
            creneaux_occupes_ens = disponibilites_enseignants.get(s_nom, set())
            if creneau_key in creneaux_occupes_ens:
                continue
            
            anti_succession_val = False
            if creneau_examen in creneaux_actifs:
                idx_c = creneaux_actifs.index(creneau_examen)
                if idx_c > 0:
                    creneau_precedent = creneaux_actifs[idx_c - 1]
                    if (d_key, creneau_precedent) in creneaux_occupes_ens:
                        anti_succession_val = True
            if anti_succession_val:
                continue

            quota = st.session_state.get(f"nb_surv_{q_val.lower()}", 3)
            current_count = row_s.iloc[0]['surveillance_attribuee']
            
            if current_count < quota:
                st.session_state.round_robin_pointer = (ptr + 1) % n_total
                return s_nom, q_val
        
        if qualite_souhaitee:
            return trouver_surveillant_round_robin(None)
            
        return None, None

    for idx, examen in planning_df.iterrows():
        date_examen = examen.get('date', None)
        creneau_examen = examen.get('Horaire', creneaux_actifs[0] if creneaux_actifs else '08h30 - 10h30')
        matiere_examen = examen.get('Enseignements', '')
        enseignant_matiere = examen.get('Enseignants', '')
        lieu_examen = examen.get('Lieu', 'S01')
        promotion_examen = examen.get('Promotion', '')
        groupe_examen = examen.get('Groupe', 'Global')
        
        if date_examen is None or pd.isna(date_examen):
            continue
            
        if isinstance(date_examen, datetime):
            d_key = date_examen.date()
        elif isinstance(date_examen, str):
            try:
                d_key = datetime.strptime(date_examen, "%Y-%m-%d").date()
            except:
                d_key = date_examen
        else:
            d_key = date_examen

        creneau_key = (d_key, creneau_examen)
        is_amphi = str(lieu_examen).strip().upper().startswith('A')
        nb_surv_requis = st.session_state.get('nb_surv_par_amphi', 3) if is_amphi else st.session_state.get('nb_surv_par_salle', 2)
            
        liste_surveillants = []
        cle_multi_lieux = (d_key, creneau_examen, matiere_examen, enseignant_matiere)
        
        if enseignant_matiere and str(enseignant_matiere) not in ['nan', '', 'None']:
            if cle_multi_lieux not in charges_affectees_creneau:
                ens_info = surveillants[surveillants['nom'] == enseignant_matiere]
                if not ens_info.empty:
                    nom_ens = ens_info.iloc[0]['nom']
                    qualite_ens = ens_info.iloc[0]['qualite']
                    
                    creneaux_occupes_ens = disponibilites_enseignants.get(nom_ens, set())
                    chevauchement = creneau_key in creneaux_occupes_ens
                    
                    anti_succession = False
                    if creneau_examen in creneaux_actifs:
                        idx_c = creneaux_actifs.index(creneau_examen)
                        if idx_c > 0:
                            creneau_precedent = creneaux_actifs[idx_c - 1]
                            if (d_key, creneau_precedent) in creneaux_occupes_ens:
                                anti_succession = True

                    if not chevauchement and not anti_succession and nom_ens not in exclus:
                        quota_key = f"nb_surv_{qualite_ens.lower()}"
                        quota = st.session_state.get(quota_key, 3)
                        current_count = surveillants.loc[surveillants['nom'] == nom_ens, 'surveillance_attribuee'].values[0]
                        if current_count < quota:
                            liste_surveillants.append({'nom': nom_ens, 'qualite': qualite_ens, 'priorite': 'Charge de matiere'})
                            disponibilites_enseignants[nom_ens].add(creneau_key)
                            surveillants.loc[surveillants['nom'] == nom_ens, 'surveillance_attribuee'] += 1
                            charges_affectees_creneau.add(cle_multi_lieux)

        if nb_surv_requis == 3:
            target_vacataires = 1
            target_permanents = 2
        elif nb_surv_requis == 2:
            target_vacataires = 0
            target_permanents = 2
        else:
            target_vacataires = 0
            target_permanents = nb_surv_requis

        while len(liste_surveillants) < nb_surv_requis:
            current_vacs = sum(1 for s in liste_surveillants if s['qualite'] == 'Vacataire')
            current_perms = sum(1 for s in liste_surveillants if s['qualite'] == 'Permanent')
            
            q_needed = None
            if current_vacs < target_vacataires:
                q_needed = 'Vacataire'
            elif current_perms < target_permanents:
                q_needed = 'Permanent'
                
            s_nom, q_val = trouver_surveillant_round_robin(q_needed)
            if s_nom:
                liste_surveillants.append({'nom': s_nom, 'qualite': q_val, 'priorite': 'Surveillant'})
                disponibilites_enseignants[s_nom].add(creneau_key)
                surveillants.loc[surveillants['nom'] == s_nom, 'surveillance_attribuee'] += 1
            else:
                break
                
        attributions.append({
            'date': date_examen, 'creneau': creneau_examen, 'matiere': matiere_examen,
            'enseignant': enseignant_matiere, 'promotion': promotion_examen, 'groupe': groupe_examen,
            'lieu': lieu_examen, 'surveillants': [s['nom'] for s in liste_surveillants], 'details_surveillants': liste_surveillants
        })
        
    attributions = sorted(attributions, key=lambda x: (x.get('date', datetime.min if 'datetime' in globals() else str(x.get('date'))), creneaux_actifs.index(x.get('creneau')) if x.get('creneau') in creneaux_actifs else 0, x.get('promotion', '')))
    
    surveillants['surveillance_attribuee'] = 0
    for attr in attributions:
        for s_nom in attr['surveillants']:
            mask = surveillants['nom'] == s_nom
            if mask.any():
                current_val = surveillants.loc[mask, 'surveillance_attribuee'].values[0]
                surveillants.loc[mask, 'surveillance_attribuee'] = current_val + 1
    
    st.session_state.disponibilites_enseignants = disponibilites_enseignants
    return attributions, surveillants

def construire_grille_edt(attributions, creneaux_liste):
    if not attributions:
        return None, None, None
    grille = {}
    jours_ordre = []
    for attr in attributions:
        date_val = attr.get('date', None)
        if date_val is None: continue
        if isinstance(date_val, str):
            try: date_val = datetime.strptime(date_val, "%Y-%m-%d").date()
            except: continue
        elif isinstance(date_val, datetime): date_val = date_val.date()
        jour_nom = JOURS_FR.get(date_val.strftime('%A'), date_val.strftime('%A'))
        date_str = date_val.strftime('%d/%m/%Y')
        cle_jour = f"{jour_nom}\n{date_str}"
        if cle_jour not in grille:
            grille[cle_jour] = {}
            jours_ordre.append(cle_jour)
        creneau = attr.get('creneau', CRENEAUX_DEFAUT[0])
        if creneau not in grille[cle_jour]: grille[cle_jour][creneau] = []
        survs = attr.get('details_surveillants', [])
        surv_text = "\n".join([f"• {s['nom']} ({'Vac' if s['qualite'] == 'Vacataire' else 'Perm' if s['qualite'] == 'Permanent' else s['qualite']})" for s in survs])
        grp_str = f" [Grp: {attr.get('groupe', 'Global')}]" if attr.get('groupe') and attr.get('groupe') != 'Global' else ""
        grille[cle_jour][creneau].append({
            'matiere': attr.get('matiere', ''), 
            'enseignant': attr.get('enseignant', ''),
            'lieu': attr.get('lieu', ''), 
            'surveillants': surv_text, 
            'promotion': f"{attr.get('promotion', '')}{grp_str}",
            'creneau': creneau,
            'date': date_str
        })
    creneaux_utilises = st.session_state.get('creneaux_actifs', CRENEAUX_DEFAUT)
    data = []
    for creneau in creneaux_utilises:
        row = {'Creneau': creneau}
        for jour in jours_ordre:
            exams = grille.get(jour, {}).get(creneau, [])
            if exams:
                cellules = []
                for ex in exams:
                    cell_text = f"📖 {ex['matiere']}\n👤 Promotion: {ex['promotion']}\n👤 Chargé: {ex['enseignant']}\n🏫 {ex['lieu']}\n👮\n{ex['surveillants']}"
                    cellules.append(cell_text)
                row[jour] = "\n---\n".join(cellules)
            else:
                row[jour] = ""
        data.append(row)
    df_grille = pd.DataFrame(data)
    return df_grille, jours_ordre, grille

def generer_excel_edt(df_grille, promotion):
    wb = Workbook()
    ws = wb.active
    ws.title = nettoyer_nom_feuille(f"EDT {promotion}")
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    creneau_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    creneau_font = Font(bold=True, size=10)
    cell_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    cell_font = Font(size=9)
    thin_border = Border(left=Side(style='thin', color='90CAF9'), right=Side(style='thin', color='90CAF9'),
                         top=Side(style='thin', color='90CAF9'), bottom=Side(style='thin', color='90CAF9'))
    headers = ['Creneau'] + [c for c in df_grille.columns if c != 'Creneau']

    # En-têtes
    for c_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Données
    for r_idx, row in df_grille.iterrows():
        for c_idx, col_name in enumerate(headers, 1):
            val = row.get(col_name, '')
            cell = ws.cell(row=r_idx + 2, column=c_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if col_name == 'Creneau':
                cell.fill = creneau_fill
                cell.font = creneau_font
            else:
                cell.fill = cell_fill
                cell.font = cell_font

    # Ajustement automatique de la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    lines_in_cell = str(cell.value).split('\n')
                    max_line_len = max(len(line) for line in lines_in_cell)
                    max_length = max(max_length, max_line_len)
            except:
                pass
        adjusted_width = min(max_length + 3, 60)
        ws.column_dimensions[column].width = max(adjusted_width, 12)

    # Ajustement automatique de la hauteur des lignes
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        max_lines = 1
        for cell in row:
            if cell.value:
                num_lines = str(cell.value).count('\n') + 1
                max_lines = max(max_lines, num_lines)
        ws.row_dimensions[row[0].row].height = max(max_lines * 14, 30)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generer_html_edt(df_grille, promotion):
    jours_cols = [c for c in df_grille.columns if c != 'Creneau']
    html = f"""
    <style>
        .edt-table {{ border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; font-size: 11px; text-align: center; }}
        .edt-table th {{ background-color: #1565C0; color: white; padding: 10px; text-align: center; border: 2px solid #0D47A1; font-size: 12px; }}
        .edt-table td {{ padding: 8px; border: 1px solid #90CAF9; vertical-align: middle; text-align: center; min-width: 200px; }}
        .creneau-cell {{ background-color: #E3F2FD; font-weight: bold; text-align: center; font-size: 12px; width: 120px; }}
        .exam-cell {{ background-color: #FFF8E1; text-align: center; }}
        .matiere {{ font-weight: bold; color: #1565C0; font-size: 12px; text-align: center; }}
        .promo {{ color: #00796B; font-size: 10px; font-weight: bold; text-align: center; }}
        .ens {{ color: #333; font-size: 10px; text-align: center; }}
        .lieu {{ color: #E65100; font-size: 10px; font-weight: bold; text-align: center; }}
        .surv {{ color: #2E7D32; font-size: 10px; text-align: center; }}
        .sep {{ border-top: 1px dashed #ccc; margin: 4px 0; }}
    </style>
    <h2 style="color:#1565C0; text-align:center;">{TITRE_PLATEFORME}</h2>
    <h3 style="color:#333; text-align:center;">Promotion {promotion}</h3>
    <table class="edt-table">
    """
    html += "<tr><th>Creneau / Horaire</th>"
    for jour in jours_cols:
        html += f"<th>{jour.replace(chr(10), '<br>')}</th>"
    html += "</tr>"
    for _, row in df_grille.iterrows():
        html += f"<tr><td class='creneau-cell'>{row['Creneau']}</td>"
        for jour in jours_cols:
            val = row.get(jour, '')
            if val:
                parts = val.split('\n---\n')
                cells_html = []
                for part in parts:
                    lines = part.split('\n')
                    formatted = []
                    for line in lines:
                        if line.startswith('📖 '): formatted.append(f"<div class='matiere'>{line[2:]}</div>")
                        elif line.startswith('👤 Promotion: '): formatted.append(f"<div class='promo'>Promotion: {line[14:]}</div>")
                        elif line.startswith('👤 '): formatted.append(f"<div class='ens'>{line[2:]}</div>")
                        elif line.startswith('🏫 '): formatted.append(f"<div class='lieu'>{line[2:]}</div>")
                        elif line.startswith('👮'): formatted.append(f"<div class='surv'>{line}</div>")
                        elif line.startswith('• '): formatted.append(f"<div class='surv'>{line}</div>")
                        else: formatted.append(f"<div>{line}</div>")
                    cells_html.append("".join(formatted))
                content = "<div class='sep'></div>".join(cells_html)
                html += f"<td class='exam-cell'>{content}</td>"
            else:
                html += "<td></td>"
        html += "</tr>"
    html += "</table>"
    return html

def generer_pdf_edt(attributions, promotion, creneaux_liste):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=13, textColor=colors.HexColor('#1565C0'), spaceAfter=6, alignment=1)
    subtitle_style = ParagraphStyle('SubTitle', parent=styles['Normal'], alignment=1, fontSize=10, textColor=colors.HexColor('#333333'), spaceAfter=10)
    
    elements.append(Paragraph(TITRE_PLATEFORME, title_style))
    elements.append(Paragraph(f"Promotion {promotion}", subtitle_style))
    elements.append(Spacer(1, 0.2*cm))
    
    creneaux_actifs = st.session_state.get('creneaux_actifs', CRENEAUX_DEFAUT)
    df_grille, jours_ordre, _ = construire_grille_edt(attributions, creneaux_actifs)
    if df_grille is None:
        elements.append(Paragraph("Aucune donnée", styles['Normal']))
        doc.build(elements)
        buffer.seek(0)
        return buffer
        
    jours_cols = [c for c in df_grille.columns if c != 'Creneau']
    
    header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=8, leading=10, textColor=colors.whitesmoke, alignment=1, fontName='Helvetica-Bold')
    creneau_header_style = ParagraphStyle('CreneauHeader', parent=header_style, fontSize=8, leading=10)
    
    table_data = [[Paragraph('Creneau / Horaire', creneau_header_style)] + [Paragraph(j.replace('\n', '<br/>'), header_style) for j in jours_cols]]
    
    cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=7, leading=9, alignment=1, textColor=colors.HexColor('#333333'))
    creneau_cell_style = ParagraphStyle('CreneauCell', parent=cell_style, fontName='Helvetica-Bold', textColor=colors.HexColor('#1565C0'))
    
    for _, row in df_grille.iterrows():
        row_data = [Paragraph(str(row['Creneau']), creneau_cell_style)]
        for jour in jours_cols:
            val = row.get(jour, '')
            if val:
                parts = val.split('\n---\n')
                formatted_parts = []
                for part in parts:
                    lines = part.split('\n')
                    f_lines = []
                    for line in lines:
                        if line.startswith('📖 '):
                            f_lines.append(f"<b>{line}</b>")
                        elif line.startswith('👤 Promotion: '):
                            f_lines.append(f"<font color='#00796B'><b>Promotion: {line[14:]}</b></font>")
                        elif line.startswith('👤 '):
                            f_lines.append(f"{line.replace('👤 ', '')}")
                        elif line.startswith('🏫 '):
                            f_lines.append(f"<font color='#E65100'><b>{line}</b></font>")
                        elif line.startswith('• '):
                            f_lines.append(f"<font color='#2E7D32'>{line}</font>")
                        elif line == '👮':
                            continue
                        else:
                            f_lines.append(line)
                    formatted_parts.append("<br/>".join(f_lines))
                cell_content = "<br/><br/>".join(formatted_parts)
                row_data.append(Paragraph(cell_content, cell_style))
            else:
                row_data.append(Paragraph("", cell_style))
        table_data.append(row_data)
        
    available_width = 27.7 * cm
    col_widths = [3 * cm] + [(available_width - 3 * cm) / len(jours_cols)] * len(jours_cols)
    
    table = Table(table_data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#90CAF9')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#E3F2FD'), colors.HexColor('#FFFFFF')]),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer

def generer_tableau_html(attributions, creneaux_utilises):
    if not attributions: return "<p style='text-align:center;'>Aucune attribution</p>"
    planning_par_jour = {}
    for attr in attributions:
        date_val = attr.get('date', None)
        if date_val is None: continue
        if isinstance(date_val, str):
            try: date_val = datetime.strptime(date_val, "%Y-%m-%d").date()
            except: continue
        elif isinstance(date_val, datetime): date_val = date_val.date()
        date_str = date_val.strftime('%d/%m/%Y')
        jour_fr = JOURS_FR.get(date_val.strftime('%A'), date_val.strftime('%A'))
        cle = f"{jour_fr} {date_str}"
        if cle not in planning_par_jour: planning_par_jour[cle] = {}
        creneau = attr.get('creneau', CRENEAUX_DEFAUT[0])
        if creneau not in planning_par_jour[cle]: planning_par_jour[cle][creneau] = []
        planning_par_jour[cle][creneau].append(attr)
    html = f"""
    <style>
        .planning-table {{ border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; font-size: 11px; text-align: center; }}
        .planning-table th {{ background-color: #1565C0; color: white; padding: 8px; text-align: center; border: 2px solid #0D47A1; }}
        .planning-table td {{ padding: 6px; border: 1px solid #90CAF9; vertical-align: middle; text-align: center; }}
        .creneau-cell {{ background-color: #E3F2FD; font-weight: bold; text-align: center; width: 120px; }}
        .examen-cell {{ background-color: #FFF8E1; margin: 2px auto; padding: 5px; border-radius: 3px; border-left: 3px solid #FFA000; font-size: 10px; text-align: center; }}
    </style>
    <h3 style="color:#1565C0; text-align:center;">{TITRE_PLATEFORME}</h3>
    <table class="planning-table">
    """
    jours = sorted(planning_par_jour.keys())
    html += "<tr><th>Creneau / Horaire</th>"
    for jour in jours: html += f"<th>{jour}</th>"
    html += "</tr>"
    creneaux_a_afficher = st.session_state.get('creneaux_actifs', CRENEAUX_DEFAUT)
    for creneau in creneaux_a_afficher:
        html += f"<tr><td class='creneau-cell'>{creneau}</td>"
        for jour in jours:
            html += "<td>"
            if creneau in planning_par_jour.get(jour, {}):
                for examen in planning_par_jour[jour][creneau]:
                    survs = examen.get('details_surveillants', [])
                    surv_html = "<br>".join([f"<span>{s['nom']} ({'Vac' if s['qualite'] == 'Vacataire' else 'Perm' if s['qualite'] == 'Permanent' else s['qualite']}{'*' if s.get('priorite') == 'Charge de matiere' else ''})</span>" for s in survs])
                    grp_info = f" | Grp: {examen.get('groupe', 'Global')}" if examen.get('groupe') and examen.get('groupe') != 'Global' else ""
                    html += f"<div class='examen-cell'><strong>{examen.get('matiere', '')}</strong><br><small>Promo: {examen.get('promotion', '')}{grp_info} | Lieu: {examen.get('lieu', '')}</small><br><small>Chargé: {examen.get('enseignant', '')}</small><br><small>{surv_html}</small></div>"
            html += "</td>"
        html += "</tr>"
    html += "</table>"
    return html

def generer_excel_colore(attributions):
    wb = Workbook()
    ws = wb.active
    ws.title = nettoyer_nom_feuille("Planning Global")
    data = []
    for attr in attributions:
        date_val = attr.get('date', None)
        if hasattr(date_val, 'strftime'):
            date_str = date_val.strftime('%d/%m/%Y')
            jour = JOURS_FR.get(date_val.strftime('%A'), date_val.strftime('%A'))
        else:
            date_str = str(date_val)
            jour = ''
        surv_str = ", ".join([f"{s['nom']} ({'Vac' if s['qualite'] == 'Vacataire' else 'Perm' if s['qualite'] == 'Permanent' else s['qualite']})" for s in attr.get('details_surveillants', [])])
        data.append({
            'Enseignements': attr.get('matiere', ''), 
            'Code': f"CODE-{abs(hash(attr.get('matiere', ''))) % 9000 + 1000}", 
            'Enseignants': attr.get('enseignant', ''), 
            'Horaire': attr.get('creneau', CRENEAUX_DEFAUT[0]), 
            'Jours': jour, 
            'Lieu': attr.get('lieu', ''), 
            'Promotion': attr.get('promotion', ''),
            'Groupe': attr.get('groupe', 'Global'),
            'Date': date_str,
            'Surveillants': surv_str
        })
    df = pd.DataFrame(data)
    cols_order = ['Enseignements', 'Code', 'Enseignants', 'Horaire', 'Jours', 'Lieu', 'Promotion', 'Groupe', 'Date', 'Surveillants']
    df = df[[c for c in cols_order if c in df.columns]]
    
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.border = Border(left=Side(style='thin', color='90CAF9'), right=Side(style='thin', color='90CAF9'), top=Side(style='thin', color='90CAF9'), bottom=Side(style='thin', color='90CAF9'))
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value: max_length = max(max_length, len(str(cell.value)))
            except: pass
        ws.column_dimensions[column].width = min(max(max_length + 2, 15), 50)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generer_pdf(attributions):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=12, textColor=colors.HexColor('#1565C0'), spaceAfter=10, alignment=1)
    elements.append(Paragraph(TITRE_PLATEFORME, title_style))
    elements.append(Paragraph("PLANNING CHRONOLOGIQUE DES SURVEILLANCES", styles['Heading2']))
    elements.append(Spacer(1, 0.3*cm))
    table_data = [['Enseignements', 'Code', 'Enseignants', 'Horaire', 'Jours', 'Lieu', 'Promotion', 'Groupe', 'Date', 'Surveillants']]
    for attr in attributions:
        date_val = attr.get('date', None)
        if hasattr(date_val, 'strftime'):
            date_str = date_val.strftime('%d/%m/%Y')
            jour = JOURS_FR.get(date_val.strftime('%A'), date_val.strftime('%A'))
        else:
            date_str = str(date_val)
            jour = ''
        surv_str = ", ".join([f"{s['nom']} ({'Vac' if s['qualite'] == 'Vacataire' else 'Perm' if s['qualite'] == 'Permanent' else s['qualite']})" for s in attr.get('details_surveillants', [])])
        table_data.append([
            attr.get('matiere', ''), 
            f"CODE-{abs(hash(attr.get('matiere', ''))) % 9000 + 1000}", 
            attr.get('enseignant', ''), 
            attr.get('creneau', CRENEAUX_DEFAUT[0]), 
            jour, 
            attr.get('lieu', ''), 
            attr.get('promotion', ''), 
            attr.get('groupe', 'Global'),
            date_str, 
            surv_str
        ])
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#90CAF9')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#E3F2FD'), colors.HexColor('#FFFFFF')]),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm
import os
import re

# Titre de la plateforme selon les consignes
TITRE_PLATEFORME = "Plateforme de gestion des EDTs-S1-2027-Département d'Électrotechnique-Faculté de génie électrique-UDL-SBA"

st.set_page_config(page_title=TITRE_PLATEFORME, page_icon="📋", layout="wide", initial_sidebar_state="expanded")

def envoyer_code_otp(nom_etud, email_dest, code_otp):
    """Envoie le code OTP à l'enseignant via SMTP."""
    body = f"Bonjour {nom_etud},\n\nVotre code d'accès à la Plateforme de Suivi d'Assiduité est : {code_otp}\n\nCe code est valable 10 minutes.\n\ndépartement d'Électrotechnique - FGE/UDL-SBA"
    
    msg = MIMEText(body)
    msg["Subject"] = "Code d'accès - Plateforme Assiduité"
    msg["From"] = "chef.department.elt.fge@gmail.com"
    msg["To"] = str(email_dest).strip()
    
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        # Utilisation des paramètres fournis[cite: 1]
        server.login("chef.department.elt.fge@gmail.com", "gkzs pdza yodb icvd")
        server.sendmail(msg["From"], msg["To"], msg.as_string())
        server.quit()
        return True
    except Exception as e:
        st.error(f"Erreur d'envoi OTP : {e}")
        return False

# --- Fonction d'envoi d'e-mail (EDT) ---
def envoyer_email_edt(destinataire, sujet, corps, fichier_buffer, nom_fichier_piece):
    """Envoie un e-mail avec l'EDT en pièce jointe."""
    smtp_user = "chef.department.elt.fge@gmail.com"
    smtp_password = "gkzs pdza yodb icvd"
    
    msg = MIMEMultipart()
    msg['From'] = smtp_user
    msg['To'] = destinataire
    msg['Subject'] = sujet
    msg.attach(MIMEText(corps, 'plain', 'utf-8'))

    if fichier_buffer and nom_fichier_piece:
        fichier_buffer.seek(0)
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(fichier_buffer.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{nom_fichier_piece}"')
        msg.attach(part)

    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(smtp_user, smtp_password)
        server.sendmail(smtp_user, destinataire, msg.as_string())
        server.quit()
        return True, "E-mail envoyé avec succès !"
    except Exception as e:
        return False, f"Erreur lors de l'envoi : {str(e)}"
    # Création du message
    msg = MIMEMultipart()
    msg['From'] = smtp_user
    msg['To'] = destinataire
    msg['Subject'] = sujet

    # Ajout du corps du message
    msg.attach(MIMEText(corps, 'plain', 'utf-8'))

    # Gestion de la pièce jointe à partir du buffer (BytesIO)
    if fichier_buffer is not None and nom_fichier_piece:
        try:
            fichier_buffer.seek(0)
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(fichier_buffer.read())
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename="{nom_fichier_piece}"'
            )
            msg.attach(part)
        except Exception as e:
            print(f"Erreur lors de l'attachement du fichier : {e}")
            return False

    # Connexion et envoi via SMTP
    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()  # Sécurisation de la connexion
        if smtp_user and smtp_password:
            server.login(smtp_user, smtp_password)
        server.sendmail(smtp_user, destinataire, msg.as_string())
        server.quit()
        return True
    except Exception as e:
        print(f"Erreur lors de l'envoi de l'e-mail : {e}")
        return False

    if not smtp_user or not smtp_password:
        return False, "Veuillez configurer vos paramètres SMTP dans la barre latérale."

    try:
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To'] = destinataire
        msg['Subject'] = sujet

        msg.attach(MIMEText(corps, 'plain', 'utf-8'))

        if fichier_buffer and nom_fichier_piece:
            fichier_buffer.seek(0)
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(fichier_buffer.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{nom_fichier_piece}"')
            msg.attach(part)

        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_password)
        server.sendmail(smtp_user, destinataire, msg.as_string())
        server.quit()
        return True, "E-mail envoyé avec succès !"
    except Exception as e:
        return False, f"Erreur lors de l'envoi : {str(e)}"



def generer_excel_toutes_promotions():
    """Génère un fichier Excel multi-feuilles avec la grille EDT de chaque promotion."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    creneau_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    creneau_font = Font(bold=True, size=10)
    cell_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    cell_font = Font(size=9)
    thin_border = Border(left=Side(style='thin', color='90CAF9'), right=Side(style='thin', color='90CAF9'),
                         top=Side(style='thin', color='90CAF9'), bottom=Side(style='thin', color='90CAF9'))

    for promo in st.session_state.promotions_list:
        attr_promo = [a for a in st.session_state.surveillance_df if str(a.get('promotion', '')).strip() == str(promo).strip()]
        if not attr_promo:
            continue
        df_grille, _, _ = construire_grille_edt(attr_promo, st.session_state.creneaux_actifs)
        if df_grille is None or df_grille.empty:
            continue

        ws = wb.create_sheet(title=nettoyer_nom_feuille(f"EDT {promo}"))
        headers = ['Creneau'] + [c for c in df_grille.columns if c != 'Creneau']

        for c_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        for r_idx, row in df_grille.iterrows():
            for c_idx, col_name in enumerate(headers, 1):
                val = row.get(col_name, '')
                cell = ws.cell(row=r_idx + 2, column=c_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if col_name == 'Creneau':
                    cell.fill = creneau_fill
                    cell.font = creneau_font
                else:
                    cell.fill = cell_fill
                    cell.font = cell_font

        # Ajustement automatique de la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        lines_in_cell = str(cell.value).split('\n')
                        max_line_len = max(len(line) for line in lines_in_cell)
                        max_length = max(max_length, max_line_len)
                except:
                    pass
            adjusted_width = min(max_length + 3, 60)
            ws.column_dimensions[column].width = max(adjusted_width, 12)

        # Ajustement automatique de la hauteur des lignes
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            max_lines = 1
            for cell in row:
                if cell.value:
                    num_lines = str(cell.value).count('\n') + 1
                    max_lines = max(max_lines, num_lines)
            ws.row_dimensions[row[0].row].height = max(max_lines * 14, 30)

    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet(title="Vide")
        ws.cell(row=1, column=1, value="Aucune donnée disponible")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generer_pdf_toutes_promotions():
    """Génère un PDF multi-pages avec la grille EDT de chaque promotion."""
    from reportlab.platypus import PageBreak
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=13, textColor=colors.HexColor('#1565C0'), spaceAfter=6, alignment=1)
    subtitle_style = ParagraphStyle('SubTitle', parent=styles['Normal'], alignment=1, fontSize=10, textColor=colors.HexColor('#333333'), spaceAfter=10)

    first_page = True
    for promo in st.session_state.promotions_list:
        attr_promo = [a for a in st.session_state.surveillance_df if str(a.get('promotion', '')).strip() == str(promo).strip()]
        if not attr_promo:
            continue
        df_grille, jours_ordre, _ = construire_grille_edt(attr_promo, st.session_state.creneaux_actifs)
        if df_grille is None or df_grille.empty:
            continue

        if not first_page:
            elements.append(PageBreak())
        first_page = False

        elements.append(Paragraph(TITRE_PLATEFORME, title_style))
        elements.append(Paragraph(f"Promotion {promo}", subtitle_style))
        elements.append(Spacer(1, 0.2*cm))

        jours_cols = [c for c in df_grille.columns if c != 'Creneau']
        header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=8, leading=10, textColor=colors.whitesmoke, alignment=1, fontName='Helvetica-Bold')
        creneau_header_style = ParagraphStyle('CreneauHeader', parent=header_style, fontSize=8, leading=10)
        table_data = [[Paragraph('Creneau / Horaire', creneau_header_style)] + [Paragraph(j.replace('\n', '<br/>'), header_style) for j in jours_cols]]
        cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=7, leading=9, alignment=1, textColor=colors.HexColor('#333333'))
        creneau_cell_style = ParagraphStyle('CreneauCell', parent=cell_style, fontName='Helvetica-Bold', textColor=colors.HexColor('#1565C0'))

        for _, row in df_grille.iterrows():
            row_data = [Paragraph(str(row['Creneau']), creneau_cell_style)]
            for jour in jours_cols:
                val = row.get(jour, '')
                if val:
                    parts = val.split('\n---\n')
                    formatted_parts = []
                    for part in parts:
                        lines_pdf = part.split('\n')
                        f_lines = []
                        for line in lines_pdf:
                            if line.startswith('📖 '): f_lines.append(f"<b>{line}</b>")
                            elif line.startswith('👤 Promotion: '): f_lines.append(f"<font color='#00796B'><b>Promotion: {line[14:]}</b></font>")
                            elif line.startswith('👤 '): f_lines.append(f"{line.replace('👤 ', '')}")
                            elif line.startswith('🏫 '): f_lines.append(f"<font color='#E65100'><b>{line}</b></font>")
                            elif line.startswith('• '): f_lines.append(f"<font color='#2E7D32'>{line}</font>")
                            elif line == '👮': continue
                            else: f_lines.append(line)
                        formatted_parts.append("<br/>".join(f_lines))
                    cell_content = "<br/><br/>".join(formatted_parts)
                    row_data.append(Paragraph(cell_content, cell_style))
                else:
                    row_data.append(Paragraph("", cell_style))
            table_data.append(row_data)

        available_width = 27.7 * cm
        col_widths = [3 * cm] + [(available_width - 3 * cm) / len(jours_cols)] * len(jours_cols)
        table = Table(table_data, repeatRows=1, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#90CAF9')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#E3F2FD'), colors.HexColor('#FFFFFF')]),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.5*cm))

    if not elements:
        elements.append(Paragraph("Aucune donnée", styles['Normal']))
    doc.build(elements)
    buffer.seek(0)
    return buffer


def main():
    init_session_state()
    charger_persistence()
    st.markdown(f'<div class="main-header">{TITRE_PLATEFORME}</div>', unsafe_allow_html=True)

    if not st.session_state.data_loaded:
        with st.spinner("Chargement du fichier source..."):
            result, error = charger_fichier_source_auto()
            if result is not None:
                st.session_state.enseignants_df = result['enseignants']
                st.session_state.examens_df = result['examens']
                st.session_state.promotions_list = result['promotions']
                st.session_state.permanents_list = result['permanents']
                st.session_state.vacataires_list = result['vacataires']
                st.session_state.all_enseignants_list = result['all_enseignants']
                st.session_state.data_loaded = True
                if result['promotions']: st.session_state.promo_selected = result['promotions'][0]
                st.success(f"✅ Chargé: {result['sheet_used']} | {len(result['enseignants'])} ens. ({len(result['vacataires'])} vacataires) | {len(result['examens'])} cours | Promos: {', '.join(result['promotions'])}")
            else:
                st.error(f"❌ {error}")

    with st.sidebar:
        st.markdown("## ⚙️ Configuration")
        if st.session_state.data_loaded:
            st.markdown("### 📁 Source")
            st.success(f"{FICHIER_SOURCE} chargé")
            st.markdown(f"- Enseignants: {len(st.session_state.enseignants_df)}")
            st.markdown(f"- Vacataires détectés: {len(st.session_state.vacataires_list)}")
            st.markdown(f"- **Cours uniquement**: {len(st.session_state.examens_df)}")
            st.markdown(f"- Promotions: {', '.join(st.session_state.promotions_list)}")
            if st.button("🔄 Recharger", key="btn_reload"):
                st.session_state.data_loaded = False
                st.rerun()
        else:
            st.warning("Fichier non chargé")
            fu = st.file_uploader("Charger manuellement", type=['xlsx', 'xls'], key="manual_up")
            if fu is not None:
                with open(FICHIER_SOURCE, "wb") as f: f.write(fu.getvalue())
                st.session_state.data_loaded = False
                st.rerun()
        st.markdown("---")
        st.markdown("### ⏰ Créneaux Horaires Actifs")
        creneaux_choisis = st.multiselect(
            "Sélectionner 1, 2 ou 3 créneaux",
            CRENEAUX_DEFAUT,
            default=st.session_state.creneaux_actifs,
            key="w_creneaux_actifs_select"
        )
        if creneaux_choisis:
            st.session_state.creneaux_actifs = creneaux_choisis
        else:
            st.warning("Veuillez sélectionner au moins un créneau.")
            st.session_state.creneaux_actifs = CRENEAUX_DEFAUT

        st.markdown("---")
        st.markdown("### 📊 Quotas de Surveillance")
        st.session_state.nb_surv_permanent = st.number_input("Permanent", 0, 20, st.session_state.nb_surv_permanent, key="w_qp")
        st.session_state.nb_surv_vacataire = st.number_input("Vacataire", 0, 20, st.session_state.nb_surv_vacataire, key="w_qv")
        st.session_state.nb_surv_autre = st.number_input("Autre", 0, 20, st.session_state.nb_surv_autre, key="w_qa")
        
        st.markdown("#### 🏛️ Quotas par Type de Lieu")
        st.session_state.nb_surv_par_salle = st.number_input("Surv. par Salle", 1, 5, st.session_state.nb_surv_par_salle, key="w_ns")
        st.session_state.nb_surv_par_amphi = st.number_input("Surv. par Amphi", 1, 5, st.session_state.nb_surv_par_amphi, key="w_na")
        st.markdown("<small style='color: #666;'>Règle active : 1 vacataire & 2 permanents pour 3 surveillants (Amphi) | 2 permanents pour 2 surveillants (Salle).</small>", unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("### 📅 Période & Cadence")
        st.session_state.date_debut_val = st.date_input("Date début", st.session_state.date_debut_val, key="w_dd")
        st.session_state.date_fin_val = st.date_input("Date fin des examens", st.session_state.date_fin_val, key="w_df")
        st.session_state.nb_par_jour = st.number_input("Nombre d'examens max par jour", min_value=1, max_value=3, value=st.session_state.nb_par_jour, key="w_nb_jour")
        
        st.markdown("### 🎉 Jours Fériés (Calendrier)")
        if 'jours_feries_list' not in st.session_state:
            st.session_state.jours_feries_list = []
            
        col_pick, col_add = st.columns([2, 1])
        with col_pick:
            date_ferie_choisie = st.date_input("Choisir un férié", key="picker_ferie_unique")
        with col_add:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("Ajouter", key="btn_add_ferie"):
                if date_ferie_choisie not in st.session_state.jours_feries_list:
                    st.session_state.jours_feries_list.append(date_ferie_choisie)
                    st.session_state.jours_feries = sorted(st.session_state.jours_feries_list)
                    st.rerun()
                    
        if st.session_state.jours_feries_list:
            st.write("Jours fériés enregistrés :")
            for jf_item in list(st.session_state.jours_feries_list):
                c_jf1, c_jf2 = st.columns([3, 1])
                with c_jf1:
                    st.text(jf_item.strftime("%d/%m/%Y"))
                with c_jf2:
                    if st.button("❌", key=f"del_ferie_{jf_item}"):
                        st.session_state.jours_feries_list.remove(jf_item)
                        st.session_state.jours_feries = sorted(st.session_state.jours_feries_list)
                        st.rerun()

        st.markdown("---")
        st.info("💡 Vendredi et Samedi exclus. Dimanche est travaillable.")

    tabs = st.tabs(["🏠 Accueil", "👥 Enseignants", "📚 Planning par Promotion", "🎯 Attributions", "📅 EDT par Promotion", "📂 Répertoire des EDTs", "📊 Export Global", "📤 Envoyer l'EDT par E-mail (Surveillance)"])

    with tabs[0]:
        st.markdown(f"""
        <div class="info-box">
            <h3>{TITRE_PLATEFORME}</h3>
            <p><strong>Gestion des plannings d'examens et surveillances</strong> | Filtre: <b>Cours uniquement</b></p>
            <ul>
                <li>📁 Chargement automatique depuis <code>{FICHIER_SOURCE}</code></li>
                <li>📚 Uniquement les enseignements commençant par <b>Cours-</b></li>
                <li>⏰ <b>Créneaux par défaut intégrés en permanence dans la colonne Horaire</b></li>
                <li>🎯 Quotas différenciés : <b>Surv. par Salle (2 permanents)</b> vs <b>Surv. par Amphi (1 vacataire & 2 permanents)</b></li>
                <li>🔀 <b>Gestion du fractionnement des promotions</b> (Sous-groupes / Sections)</li>
                <li>🛠️ <b>Balayage séquentiel (Round-Robin), anti-succession et prévention des chevauchements</b></li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
        if st.session_state.data_loaded:
            c1, c2, c3, c4 = st.columns(4)
            with c1: st.metric("Enseignants", len(st.session_state.enseignants_df))
            with c2: st.metric("Vacataires", len(st.session_state.vacataires_list))
            with c3: st.metric("Cours filtrés", len(st.session_state.examens_df))
            with c4: st.metric("Promotions", len(st.session_state.promotions_list))

    with tabs[1]:
        st.markdown('<div class="sub-header">Gestion des Enseignants et Qualité</div>', unsafe_allow_html=True)
        if st.session_state.data_loaded:
            df_ens = st.session_state.enseignants_df
            all_ens = st.session_state.all_enseignants_list
            
            st.markdown("### 🔢 Afficheur Numérique du Nombre de Surveillances par Enseignant")
            if all_ens:
                ens_selectionne = st.selectbox("Sélectionner un enseignant dans la liste déroulante :", sorted(all_ens), key="select_ens_surv_display")
                if ens_selectionne:
                    # Récupérer les détails assignés
                    details_assigns = []
                    if st.session_state.surveillance_df is not None:
                        for attr in st.session_state.surveillance_df:
                            survs = [s['nom'] for s in attr.get('details_surveillants', [])]
                            if ens_selectionne in survs or attr.get('enseignant') == ens_selectionne:
                                details_assigns.append(attr)
                    
                    # ✅ CORRECTION: Compter correctement = nombre exact de lignes dans le tableau
                    nb_surv_actuel = len(details_assigns)
                                
                    col_m1, col_m2 = st.columns(2)
                    with col_m1:
                        st.metric(label=f"Total Surveillances pour : {ens_selectionne}", value=nb_surv_actuel)
                    with col_m2:
                        qualite_ens_val = df_ens[df_ens['nom'] == ens_selectionne]['qualite'].values[0] if not df_ens[df_ens['nom'] == ens_selectionne].empty else 'N/A'
                        st.metric(label="Qualité / Statut", value=qualite_ens_val)
                        
                    if details_assigns:
                        st.markdown(f"**Détails des examens / surveillances associés à {ens_selectionne} :** ({nb_surv_actuel} enregistrement(s))")
                        df_details_ens = pd.DataFrame([{
                            'Date': (a['date'].strftime('%d/%m/%Y') if hasattr(a.get('date'), 'strftime') else str(a.get('date'))),
                            'Horaire': a.get('creneau'),
                            'Matière': a.get('matiere'),
                            'Promotion': a.get('promotion'),
                            'Lieu': a.get('lieu'),
                            'Rôle': ('Chargé de matière' if a.get('enseignant') == ens_selectionne else 'Surveillant')
                        } for a in details_assigns])
                        st.dataframe(df_details_ens, use_container_width=True, hide_index=True)
                    else:
                        st.info(f"Aucune surveillance ou enseignement assigné pour l'instant à {ens_selectionne}.")
            
            st.markdown("---")
            exclus = st.multiselect("Sélectionner les enseignants à EXCLURE", sorted(all_ens), default=st.session_state.exclus_manuels, key="w_exclus")
            st.session_state.exclus_manuels = exclus
            if not df_ens.empty:
                for col_req in ['nom', 'qualite', 'Enseignements', 'Promotion']:
                    if col_req not in df_ens.columns:
                        df_ens[col_req] = ''
                disp_ens = df_ens[['nom', 'qualite', 'Enseignements', 'Promotion']].copy()
                if 'surveillance_attribuee' in df_ens.columns:
                    disp_ens['Surveillances attribuées'] = df_ens['surveillance_attribuee']
                disp_ens['Exclu'] = disp_ens['nom'].apply(lambda x: '❌ OUI' if x in exclus else '✅ Non')
                disp_ens = disp_ens.sort_values(by=['qualite', 'nom'], ascending=[True, True])
                st.dataframe(disp_ens, use_container_width=True, hide_index=True)
        else: st.warning("Données non chargées.")

    with tabs[2]:
        st.markdown('<div class="sub-header">Planification, Fractionnement & Cohérence par Promotion</div>', unsafe_allow_html=True)
        
        col_h1, col_h2 = st.columns([1, 4])
        with col_h1:
            if st.button("🗑️ Supprimer l'historique", type="secondary", key="btn_suppr_hist"):
                st.session_state.historique_edt = {}
                st.session_state.planning_df = None
                st.session_state.surveillance_df = None
                st.success("L'historique des EDTs a été supprimé avec succès.")
                st.rerun()
                
        if st.session_state.data_loaded and st.session_state.promotions_list:
            promo_selected = st.selectbox("📚 Sélectionner une promotion", st.session_state.promotions_list, index=0, key="w_promo_sel")
            st.session_state.promo_selected = promo_selected
            
            if promo_selected:
                st.markdown("#### 🔀 Gestion du Fractionnement / Sous-groupes")
                fractionner = st.checkbox(f"Activer le fractionnement (Sous-groupes) pour {promo_selected}", value=st.session_state.fractionnement_actif.get(promo_selected, False), key=f"chk_frac_{promo_selected}")
                st.session_state.fractionnement_actif[promo_selected] = fractionner
                
                groupes_actifs = ['Global']
                if fractionner:
                    groupes_saisis = st.text_input(f"Définir les sous-groupes (séparés par des virgules) pour {promo_selected}", value=", ".join(st.session_state.groupes_par_promo.get(promo_selected, ['Grp 1', 'Grp 2'])), key=f"txt_groupes_{promo_selected}")
                    groupes_actifs = [g.strip() for g in groupes_saisis.split(',') if g.strip()]
                    st.session_state.groupes_par_promo[promo_selected] = groupes_actifs
                    st.info(f"💡 Cohérence de fractionnement active : Les examens seront planifiés en cohérence avec les sous-groupes ({', '.join(groupes_actifs)}), évitant les conflits simultanés sur la même fraction.")
                
                col1, col2 = st.columns(2)
                with col1: salles_sel = st.multiselect("Salles", SALLES, default=['S01', 'S02', 'S03', 'S04', 'S05'], key=f"w_sal_{promo_selected}")
                with col2: amphis_sel = st.multiselect("Amphis", AMPHIS, default=['A01', 'A02', 'A03'], key=f"w_amp_{promo_selected}")
                lieux_sel = salles_sel + amphis_sel
                st.session_state.lieux_par_promo[promo_selected] = lieux_sel
                
                st.markdown("#### 🕒 Personnalisation de la Date et de l'Horaire par Matière")
                df_promo_matieres = st.session_state.examens_df[st.session_state.examens_df['Promotion'].astype(str).str.strip() == str(promo_selected).strip()]
                
                if not df_promo_matieres.empty:
                    if promo_selected not in st.session_state.horaires_par_matiere:
                        st.session_state.horaires_par_matiere[promo_selected] = {}
                    if promo_selected not in st.session_state.jours_par_matiere:
                        st.session_state.jours_par_matiere[promo_selected] = {}
                        
                    for _, row_mat in df_promo_matieres.iterrows():
                        m_nom = row_mat['Enseignements']
                        col_m1, col_m2, col_m3 = st.columns([2, 1, 1])
                        with col_m1:
                            st.text(f"📖 {m_nom} (Resp: {row_mat['Enseignants']})")
                        with col_m2:
                            date_actuelle = st.session_state.jours_par_matiere.get(promo_selected, {}).get(m_nom, None)
                            activer_date_fixe = st.checkbox(f"Date fixe - {m_nom}", value=(date_actuelle is not None), key=f"chk_date_{promo_selected}_{m_nom}")
                            if activer_date_fixe:
                                d_val = date_actuelle if isinstance(date_actuelle, date) else st.session_state.date_debut_val
                                d_choisie = st.date_input(f"Date - {m_nom}", value=d_val, key=f"d_choisie_{promo_selected}_{m_nom}")
                                st.session_state.jours_par_matiere.setdefault(promo_selected, {})[m_nom] = d_choisie
                            else:
                                if promo_selected in st.session_state.jours_par_matiere and m_nom in st.session_state.jours_par_matiere[promo_selected]:
                                    del st.session_state.jours_par_matiere[promo_selected][m_nom]
                        with col_m3:
                            options_h = ["Sélection manuelle"] + st.session_state.creneaux_actifs
                            horaire_actuel = st.session_state.horaires_par_matiere.get(promo_selected, {}).get(m_nom, "Sélection manuelle")
                            try:
                                idx_h = options_h.index(horaire_actuel)
                            except:
                                idx_h = 0
                            h_choisi = st.selectbox(f"Horaire - {m_nom}", options_h, index=idx_h, key=f"h_{promo_selected}_{m_nom}")
                            if h_choisi != "Sélection manuelle":
                                st.session_state.horaires_par_matiere.setdefault(promo_selected, {})[m_nom] = h_choisi
                            else:
                                if promo_selected in st.session_state.horaires_par_matiere and m_nom in st.session_state.horaires_par_matiere[promo_selected]:
                                    del st.session_state.horaires_par_matiere[promo_selected][m_nom]
                                
                # Vérification anti-régénération
                deja_genere = promo_selected in st.session_state.get('historique_edt', {}) and bool(st.session_state.historique_edt.get(promo_selected))
                if deja_genere:
                    st.warning(f"⚠️ Un EDT existe déjà pour la promotion **{promo_selected}**. Vous pouvez le consulter dans les onglets 📅 EDT par Promotion ou 📂 Répertoire des EDTs.")
                    force_regen = st.checkbox(f"🔄 Forcer la régénération (écraser l'EDT existant de {promo_selected})", key=f"force_regen_{promo_selected}")
                else:
                    force_regen = True

                if st.button(f"🚀 Générer l'EDT pour {promo_selected}", type="primary", key=f"btn_gen_{promo_selected}"):
                    if not force_regen:
                        st.error("❌ Régénération refusée. Cochez 'Forcer la régénération' pour écraser l'EDT existant.")
                    elif len(lieux_sel) == 0: 
                        st.error("❌ Sélectionnez au moins un lieu.")
                    else:
                        if st.session_state.planning_df is None:
                            st.session_state.planning_df = st.session_state.examens_df.copy()
                            if 'Groupe' not in st.session_state.planning_df.columns:
                                st.session_state.planning_df['Groupe'] = 'Global'
                            
                        if fractionner and groupes_actifs:
                            base_rows = []
                            for _, r in st.session_state.examens_df[st.session_state.examens_df['Promotion'].astype(str).str.strip() == str(promo_selected).strip()].iterrows():
                                for g in groupes_actifs:
                                    r_copy = r.copy()
                                    r_copy['Groupe'] = g
                                    base_rows.append(r_copy)
                            df_frac = pd.DataFrame(base_rows)
                            st.session_state.planning_df = st.session_state.planning_df[st.session_state.planning_df['Promotion'].astype(str).str.strip() != str(promo_selected).strip()]
                            st.session_state.planning_df = pd.concat([st.session_state.planning_df, df_frac], ignore_index=True)

                        ordre = st.session_state.ordre_matieres.get(promo_selected, None)
                        horaires = st.session_state.horaires_par_matiere.get(promo_selected, None)
                        jours_m = st.session_state.jours_par_matiere.get(promo_selected, None)
                        lieux_p = st.session_state.lieux_par_promo.get(promo_selected, lieux_sel)
                        
                        df_updated = generer_planning_promo(
                            st.session_state.planning_df, 
                            promo_selected, 
                            st.session_state.date_debut_val, 
                            st.session_state.date_fin_val, 
                            st.session_state.nb_par_jour, 
                            st.session_state.jours_feries, 
                            st.session_state.creneaux_actifs, 
                            lieux_p, 
                            ordre, 
                            horaires, 
                            jours_m,
                            groupes_actifs if fractionner else None
                        )
                        if df_updated is not None:
                            st.session_state.planning_df = df_updated
                            promo_subset = df_updated[df_updated['Promotion'].astype(str).str.strip() == str(promo_selected).strip()]
                            st.session_state.historique_edt[promo_selected] = promo_subset.to_dict('records')
                            st.success(f"✅ Génération avec unification des lieux et sous-groupes effectuée pour la promotion {promo_selected} !")
                            sauvegarder_persistence()

                if st.session_state.planning_df is not None:
                    st.markdown("---")
                    st.markdown(f"#### 📝 Édition directe du Planning de la promotion : {promo_selected}")
                    st.info("💡 L'éditeur bloque automatiquement toute date correspondant à un weekend (Vendredi/Samedi) ou un jour férié.")
                    
                    planning_display = st.session_state.planning_df[st.session_state.planning_df['Promotion'].astype(str).str.strip() == str(promo_selected).strip()].copy()
                    if not planning_display.empty:
                        colonnes_edition = ["Enseignements", "Groupe", "date", "Horaire", "Lieu", "Enseignants", "Promotion"]
                        for col_c in colonnes_edition:
                            if col_c not in planning_display.columns:
                                planning_display[col_c] = ""
                                
                        df_edit_result = st.data_editor(
                            planning_display[colonnes_edition],
                            key=f"editor_planning_{promo_selected}",
                            use_container_width=True,
                            hide_index=True,
                            num_rows="fixed",
                            column_config={
                                "Horaire": st.column_config.SelectboxColumn(
                                    "Horaire",
                                    help="Sélectionner le créneau horaire",
                                    options=st.session_state.creneaux_actifs,
                                    required=True
                                )
                            }
                        )
                        
                        if st.button("💾 Enregistrer les modifications du planning", type="primary", key=f"btn_save_edit_{promo_selected}"):
                            erreur_detectee = False
                            for idx_ed, row_ed in df_edit_result.iterrows():
                                new_date = row_ed['date']
                                if isinstance(new_date, str):
                                    try:
                                        new_date = datetime.strptime(new_date, "%Y-%m-%d").date()
                                    except:
                                        pass
                                
                                if not est_jour_travaille(new_date, st.session_state.jours_feries):
                                    erreur_detectee = True
                                    d_str = new_date.strftime('%d/%m/%Y') if hasattr(new_date, 'strftime') else str(new_date)
                                    st.error(f"❌ Erreur : La date {d_str} pour l'enseignement '{row_ed['Enseignements']}' est un jour férié ou un week-end (Vendredi/Samedi). Modification refusée.")
                                    break
                                    
                            if not erreur_detectee:
                                for idx_ed, row_ed in df_edit_result.iterrows():
                                    m_nom = row_ed['Enseignements']
                                    g_nom = row_ed['Groupe']
                                    new_date = row_ed['date']
                                    new_horaire = row_ed['Horaire']
                                    new_lieu = row_ed['Lieu']
                                    new_ens = row_ed['Enseignants']
                                    new_promo = row_ed['Promotion']
                                    
                                    mask = (st.session_state.planning_df['Promotion'].astype(str).str.strip() == str(promo_selected).strip()) & (st.session_state.planning_df['Enseignements'] == m_nom)
                                    if 'Groupe' in st.session_state.planning_df.columns:
                                        mask = mask & (st.session_state.planning_df['Groupe'] == g_nom)

                                    st.session_state.planning_df.loc[mask, 'date'] = new_date
                                    st.session_state.planning_df.loc[mask, 'Horaire'] = new_horaire
                                    st.session_state.planning_df.loc[mask, 'Lieu'] = new_lieu
                                    st.session_state.planning_df.loc[mask, 'Enseignants'] = new_ens
                                    st.session_state.planning_df.loc[mask, 'Promotion'] = new_promo
                                    if isinstance(new_date, (date, datetime)):
                                        st.session_state.planning_df.loc[mask, 'Jours'] = JOURS_FR.get(new_date.strftime("%A"), new_date.strftime("%A"))
                                        
                                st.success("✅ Modifications enregistrées avec succès dans le planning de la promotion !")
                                st.rerun()
                    else:
                        st.info(f"Aucun examen planifié pour {promo_selected}.")
        else: st.warning("Aucune promotion détectée.")

    with tabs[3]:
        st.markdown('<div class="sub-header">Attribution des Surveillants et Édition</div>', unsafe_allow_html=True)
        if st.session_state.planning_df is not None and st.session_state.enseignants_df is not None:
            if st.button("🎯 Attribuer les Surveillants", type="primary", key="btn_attrib"):
                with st.spinner("Attribution intelligente en cours (Round-Robin, Anti-succesion, Règle vacataires/permanents)..."):
                    attributions, ens_maj = attribuer_surveillants(st.session_state.planning_df, st.session_state.enseignants_df)
                    if attributions is not None:
                        st.session_state.surveillance_df = attributions
                        st.session_state.enseignants_df = ens_maj
                        st.success(f"✅ {len(attributions)} attributions effectuées avec succès selon les règles en vigueur !")
                        sauvegarder_persistence()
                    else: st.error("❌ Erreur lors de l'attribution.")
            if st.session_state.surveillance_df is not None:
                st.markdown("---")
                attr_data = []
                for attr in st.session_state.surveillance_df:
                    d = attr.get('date', None)
                    ds = d.strftime('%d/%m/%Y') if hasattr(d, 'strftime') else str(d) if d else ''
                    jour = JOURS_FR.get(d.strftime('%A'), d.strftime('%A')) if hasattr(d, 'strftime') else ''
                    surv_noms = ", ".join([s['nom'] for s in attr.get('details_surveillants', [])])
                    attr_data.append({
                        'Enseignements': attr.get('matiere', ''),
                        'Code': f"CODE-{abs(hash(attr.get('matiere', ''))) % 9000 + 1000}",
                        'Enseignants': attr.get('enseignant', ''),
                        'Horaire': attr.get('creneau', CRENEAUX_DEFAUT[0]),
                        'Jours': jour,
                        'Lieu': attr.get('lieu', ''),
                        'Promotion': attr.get('promotion', ''),
                        'Groupe': attr.get('groupe', 'Global'),
                        'Date': ds,
                        'Surveillants': surv_noms
                    })
                if attr_data:
                    df_attr_display = pd.DataFrame(attr_data)
                    st.markdown("#### 📝 Tableau des attributions (Édition des surveillants/lieux/horaires/dates)")
                    st.info("💡 L'éditeur d'attributions vérifie également les dates modifiées contre les weekends et jours fériés.")
                    df_edit_att = st.data_editor(
                        df_attr_display,
                        use_container_width=True,
                        hide_index=True,
                        key="editor_attributions",
                        column_config={
                            "Horaire": st.column_config.SelectboxColumn(
                                "Horaire",
                                help="Sélectionner le créneau horaire",
                                options=st.session_state.creneaux_actifs,
                                required=True
                            )
                        }
                    )
                    
                    if st.button("💾 Mettre à jour les attributions modifiées", type="primary", key="btn_save_att"):
                        erreur_attr = False
                        new_surv_df = []
                        for idx_a, row_a in df_edit_att.iterrows():
                            try:
                                parsed_date = datetime.strptime(row_a['Date'], '%d/%m/%Y').date()
                            except:
                                try:
                                    parsed_date = datetime.strptime(row_a['Date'], '%Y-%m-%d').date()
                                except:
                                    parsed_date = date(2026, 11, 1)

                            if not est_jour_travaille(parsed_date, st.session_state.jours_feries):
                                erreur_attr = True
                                st.error(f"❌ Erreur : La date '{row_a['Date']}' pour '{row_a['Enseignements']}' correspond à un jour férié ou un week-end. Modification refusée.")
                                break

                            noms_surv_list = [s.strip() for s in str(row_a['Surveillants']).split(',') if s.strip()]
                            details_surv = []
                            for n_s in noms_surv_list:
                                q_match = st.session_state.enseignants_df[st.session_state.enseignants_df['nom'].str.strip() == n_s]
                                q_val = q_match.iloc[0]['qualite'] if not q_match.empty else 'Permanent'
                                details_surv.append({'nom': n_s, 'qualite': q_val, 'priorite': 'Permanent'})
                                
                            new_surv_df.append({
                                'date': parsed_date,
                                'creneau': row_a['Horaire'],
                                'matiere': row_a['Enseignements'],
                                'enseignant': row_a['Enseignants'],
                                'promotion': row_a['Promotion'],
                                'groupe': row_a['Groupe'],
                                'lieu': row_a['Lieu'],
                                'surveillants': noms_surv_list,
                                'details_surveillants': details_surv
                            })
                            
                        if not erreur_attr:
                            st.session_state.surveillance_df = new_surv_df
                            st.success("✅ Attributions mises à jour avec succès !")
                            st.rerun()
        else: st.warning("Générez d'abord le planning.")

    with tabs[4]:
        st.markdown('<div class="sub-header">📅 EDT Chronologique en Grille par Promotion</div>', unsafe_allow_html=True)
        if st.session_state.surveillance_df is not None:
            promo_sel_choisie = st.selectbox("🎯 Choisir la promotion à afficher, télécharger et envoyer :", st.session_state.promotions_list, key="select_promo_unique_edt")
            if promo_sel_choisie:
                st.markdown(f"#### 🎓 Promotion: **{promo_sel_choisie}**")
                attr_promo = [a for a in st.session_state.surveillance_df if str(a.get('promotion', '')).strip() == str(promo_sel_choisie).strip()]
                if attr_promo:
                    df_grille, _, _ = construire_grille_edt(attr_promo, st.session_state.creneaux_actifs)
                    if df_grille is not None and not df_grille.empty:
                        st.dataframe(df_grille, use_container_width=True, hide_index=True)
                        c1, c2, c3 = st.columns(3)
                        with c1: st.download_button(f"⬇️ HTML - {promo_sel_choisie}", generer_html_edt(df_grille, promo_sel_choisie), f"EDT_{promo_sel_choisie}.html", "text/html", key=f"dl_html_{promo_sel_choisie}")
                        with c2: st.download_button(f"⬇️ Excel - {promo_sel_choisie}", generer_excel_edt(df_grille, promo_sel_choisie), f"EDT_{promo_sel_choisie}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_xlsx_{promo_sel_choisie}")
                        with c3: st.download_button(f"⬇️ PDF - {promo_sel_choisie}", generer_pdf_edt(attr_promo, promo_sel_choisie, st.session_state.creneaux_actifs), f"EDT_{promo_sel_choisie}.pdf", "application/pdf", key=f"dl_pdf_{promo_sel_choisie}")
                        
                        st.markdown("---")
                        st.markdown("#### 📤 Envoyer l'EDT par E-mail")
                        with st.form(key=f"form_email_{promo_sel_choisie}"):
                            email_destinataire = st.text_input("Adresse e-mail du destinataire (Représentant de promotion / Enseignant)")
                            sujet_email = st.text_input("Sujet", value=f"[{TITRE_PLATEFORME}] Emploi du temps - Promotion {promo_sel_choisie}")
                            corps_email = st.text_area("Message", value=f"Bonjour,\n\nVeuillez trouver ci-joint l'emploi du temps des examens pour la promotion {promo_sel_choisie}.\n\nCordialement,\nDépartement d'Électrotechnique")
                            
                            submit_email = st.form_submit_button("📨 Envoyer par E-mail")
                            if submit_email:
                                if not email_destinataire:
                                    st.error("Veuillez saisir une adresse e-mail valide.")
                                else:
                                    pdf_buffer = generer_pdf_edt(attr_promo, promo_sel_choisie, st.session_state.creneaux_actifs)
                                    succes, message = envoyer_email_edt(
                                        email_destinataire, 
                                        sujet_email, 
                                        corps_email, 
                                        pdf_buffer, 
                                        f"EDT_{promo_sel_choisie}.pdf"
                                    )
                                    if succes:
                                        st.success(f"✅ {message}")
                                    else:
                                        st.error(f"❌ {message}")
                else:
                    st.info(f"Aucune attribution trouvée pour la promotion {promo_sel_choisie}.")
        else: st.warning("⚠️ Veuillez d'abord générer les attributions.")

    with tabs[5]:
        st.markdown('<div class="sub-header">📂 Répertoire des EDTs & Téléchargements (Individuel & Groupé)</div>', unsafe_allow_html=True)
        
        nb_generes = len(st.session_state.historique_edt) if 'historique_edt' in st.session_state else 0
        total_promos_attendu = len(st.session_state.promotions_list) if st.session_state.promotions_list else 23
        
        st.markdown(f"""
        <div style="background-color: #e3f2fd; padding: 15px; border-radius: 8px; text-align: center; margin-bottom: 20px;">
            <h3 style="color: #1565c0; margin: 0;">Afficheur Numérique des EDTs</h3>
            <p style="font-size: 1.5rem; font-weight: bold; color: #0d47a1; margin: 5px 0;">
                {nb_generes} / {total_promos_attendu} EDTs générés
            </p>
            <p style="font-size: 0.9rem; color: #555; margin: 0;">Nombre d'emplois du temps enregistrés dans le répertoire sur le total de vos promotions.</p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.session_state.surveillance_df is not None and st.session_state.promotions_list:
            st.markdown("### 📥 Téléchargement Groupé de tous les EDTs")
            col_g1, col_g2 = st.columns(2)
            with col_g1:
                st.download_button("⬇️ Télécharger le Planning Global Excel", generer_excel_colore(st.session_state.surveillance_df), "Planning_Global_Toutes_Promotions.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.spreadsheet", key="btn_group_excel")
            with col_g2:
                st.download_button("⬇️ Télécharger le Planning Global PDF", generer_pdf(st.session_state.surveillance_df), "Planning_Global_Toutes_Promotions.pdf", "application/pdf", key="btn_group_pdf")
                
            st.markdown("---")
            st.markdown("### 📑 Répertoire Individuel par Promotion")
            for promo in st.session_state.promotions_list:
                attr_promo = [a for a in st.session_state.surveillance_df if str(a.get('promotion', '')).strip() == str(promo).strip()]
                if attr_promo:
                    df_g, _, _ = construire_grille_edt(attr_promo, st.session_state.creneaux_actifs)
                    if df_g is not None and not df_g.empty:
                        with st.expander(f"🎓 Promotion : {promo}"):
                            st.dataframe(df_g, use_container_width=True, hide_index=True)
                            b_ind1, b_ind2, b_ind3 = st.columns(3)
                            with b_ind1:
                                st.download_button(f"HTML - {promo}", generer_html_edt(df_g, promo), f"EDT_{promo}.html", "text/html", key=f"rep_html_{promo}")
                            with b_ind2:
                                st.download_button(f"Excel - {promo}", generer_excel_edt(df_g, promo), f"EDT_{promo}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"rep_xlsx_{promo}")
                            with b_ind3:
                                st.download_button(f"PDF - {promo}", generer_pdf_edt(attr_promo, promo, st.session_state.creneaux_actifs), f"EDT_{promo}.pdf", "application/pdf", key=f"rep_pdf_{promo}")
                else:
                    st.info(f"Promotion {promo} : Pas encore générée.")
        else:
            st.warning("⚠️ Veuillez d'abord générer les plannings et les attributions pour alimenter le répertoire.")

    with tabs[6]:
        st.markdown('<div class="sub-header">📊 Export Global — Grilles EDT par Promotion</div>', unsafe_allow_html=True)

        if st.session_state.surveillance_df is None or len(st.session_state.surveillance_df) == 0:
            st.warning("⚠️ Aucune attribution à exporter. Veuillez d'abord générer les attributions dans l'onglet 🎯 Attributions.")
        else:
            # --- TÉLÉCHARGEMENT GROUPÉ (toutes les promotions) ---
            st.markdown("### 📦 Téléchargement Groupé — Toutes les Promotions")
            col_g1, col_g2 = st.columns(2)
            with col_g1:
                st.download_button(
                    "⬇️ Télécharger TOUTES les promotions (Excel multi-feuilles)",
                    generer_excel_toutes_promotions(),
                    "Toutes_Promotions_EDT_Grille.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_all_excel"
                )
            with col_g2:
                st.download_button(
                    "⬇️ Télécharger TOUTES les promotions (PDF multi-pages)",
                    generer_pdf_toutes_promotions(),
                    "Toutes_Promotions_EDT_Grille.pdf",
                    "application/pdf",
                    key="dl_all_pdf"
                )

            st.markdown("---")
            st.markdown("### 📑 Grilles EDT par Promotion (Individuel)")

            for promo in st.session_state.promotions_list:
                attr_promo = [a for a in st.session_state.surveillance_df if str(a.get('promotion', '')).strip() == str(promo).strip()]
                if not attr_promo:
                    continue

                df_grille, jours_ordre, _ = construire_grille_edt(attr_promo, st.session_state.creneaux_actifs)
                if df_grille is None or df_grille.empty:
                    continue

                with st.container(border=True):
                    st.markdown(f"#### 🎓 Promotion : **{promo}**")
                    st.dataframe(df_grille, use_container_width=True, hide_index=True)

                    col1, col2, col3 = st.columns([1, 1, 2])
                    with col1:
                        st.download_button(
                            f"⬇️ Excel — {promo}",
                            generer_excel_edt(df_grille, promo),
                            f"EDT_{promo}_Grille.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_indiv_excel_{promo}"
                        )
                    with col2:
                        st.download_button(
                            f"⬇️ PDF — {promo}",
                            generer_pdf_edt(attr_promo, promo, st.session_state.creneaux_actifs),
                            f"EDT_{promo}_Grille.pdf",
                            "application/pdf",
                            key=f"dl_indiv_pdf_{promo}"
                        )
                    with col3:
                        st.markdown(f"<span style='color:#666; font-size:0.85rem;'>📊 {len(attr_promo)} séance(s) de surveillance | {len(jours_ordre)} jour(s)</span>", unsafe_allow_html=True)

                st.markdown("<br>", unsafe_allow_html=True)
    with tabs[7]:
        st.markdown('<div class="sub-header">📤 Envoyer l''EDT par E-mail (Surveillance)</div>', unsafe_allow_html=True)

        # Paramètres SMTP hardcodés (non modifiables via l'interface)
        smtp_server = SMTP_SERVER
        smtp_port = SMTP_PORT
        sender_email = SMTP_USER
        sender_password = SMTP_PASSWORD

        st.info(f"📧 Configuration SMTP active : {sender_email} via {smtp_server}:{smtp_port}")

        if not sender_email or not sender_password:
            st.error("⚠️ Les paramètres SMTP ne sont pas configurés. Vérifiez les constantes SMTP_USER et SMTP_PASSWORD dans le code.")

        # --- Vérification des attributions disponibles ---
        if st.session_state.surveillance_df is None or len(st.session_state.surveillance_df) == 0:
            st.warning("⚠️ Aucune attribution de surveillance disponible. Veuillez d'abord générer les attributions dans l'onglet 🎯 Attributions.")
        else:
            st.success(f"✅ {len(st.session_state.surveillance_df)} attributions de surveillance disponibles.")

            # Conversion des attributions en DataFrame exploitable
            rows = []
            for attr in st.session_state.surveillance_df:
                date_val = attr.get('date')
                date_str = ''
                jour = ''
                if hasattr(date_val, 'strftime'):
                    date_str = date_val.strftime('%d/%m/%Y')
                    jour = JOURS_FR.get(date_val.strftime('%A'), date_val.strftime('%A'))
                noms_uniques = []
                for s in attr.get('details_surveillants', []):
                    if s['nom'] not in noms_uniques:
                        noms_uniques.append(s['nom'])
                        rows.append({
                            'Enseignements': attr.get('matiere', ''),
                            'Enseignants': s['nom'],
                            'Horaire': attr.get('creneau', ''),
                            'Jours': jour,
                            'Lieu': attr.get('lieu', ''),
                            'Promotion': attr.get('promotion', ''),
                            'Groupe': attr.get('groupe', 'Global'),
                            'Date': date_str,
                            'Email': ''
                        })
            df_attributions = pd.DataFrame(rows)

            st.markdown("**📋 Aperçu des surveillances extraites des attributions :**")
            st.dataframe(df_attributions[['Enseignements', 'Enseignants', 'Horaire', 'Jours', 'Lieu', 'Promotion', 'Groupe']], use_container_width=True, hide_index=True)

            st.markdown("---")
            st.markdown("### 📧 Import du mapping Enseignant → Email")
            st.info("Importez un fichier Excel/CSV contenant les colonnes **'Enseignants'** et **'Email'** pour associer les adresses e-mails aux surveillants.")

            uploaded_mapping = st.file_uploader("Fichier de mapping Enseignant → Email (Excel/CSV)", type=["xlsx", "csv"], key="email_mapping")

            df = df_attributions.copy()
            if uploaded_mapping is not None:
                if uploaded_mapping.name.endswith('.csv'):
                    df_map = pd.read_csv(uploaded_mapping)
                else:
                    df_map = pd.read_excel(uploaded_mapping)
                if 'Enseignants' not in df_map.columns or 'Email' not in df_map.columns:
                    st.error("❌ Le fichier de mapping doit contenir obligatoirement les colonnes 'Enseignants' et 'Email'.")
                else:
                    df = df_attributions.merge(df_map[['Enseignants', 'Email']], on='Enseignants', how='left', suffixes=('', '_map'))
                    if 'Email_map' in df.columns:
                        df['Email'] = df['Email_map']
                        df = df.drop(columns=['Email_map'])
                    st.success("✅ Mapping email fusionné avec les attributions !")
            else:
                st.info("💡 Aucun fichier de mapping importé. Les e-mails sont vides. Uploadez un fichier de mapping pour activer l'envoi.")

            nb_avec_email = df['Email'].replace('', pd.NA).notna().sum() if 'Email' in df.columns else 0
            st.info(f"📊 {nb_avec_email} / {len(df)} lignes ont une adresse e-mail associée.")

            enseignants_list = df['Enseignants'].dropna().unique()

            tab_indiv, tab_groupe = st.tabs(["👤 Envoi Individuel", "👥 Envoi par Groupe (Masse)"])

            # --- 1. ENVOI INDIVIDUEL ---
            with tab_indiv:
                st.subheader("Gestion et Envoi Individuel")
                selected_prof = st.selectbox("Sélectionner un surveillant", enseignants_list, key="email_select_prof")

                df_prof = df[df['Enseignants'] == selected_prof]
                df_prof = df_prof.drop_duplicates(subset=['Enseignements', 'Horaire', 'Jours', 'Lieu', 'Promotion', 'Groupe'])

                prof_email = df_prof['Email'].iloc[0] if not df_prof.empty and 'Email' in df_prof.columns and pd.notna(df_prof['Email'].iloc[0]) and str(df_prof['Email'].iloc[0]).strip() != '' else ""

                st.text(f"E-mail associé : {prof_email if prof_email else 'Aucun e-mail trouvé'}")
                st.markdown("**Planning de surveillance affecté :**")
                st.dataframe(df_prof[['Enseignements', 'Horaire', 'Jours', 'Lieu', 'Promotion', 'Groupe']], use_container_width=True, hide_index=True)

                custom_msg = st.text_area("Message personnalisé (optionnel)", "Bonjour,\n\nVeuillez trouver ci-joint votre planning de surveillance pour les examens.\n\nCordialement,", key="email_custom_msg")

                if st.button("Envoyer l'EDT à cet enseignant", key="email_btn_indiv"):
                    if not prof_email:
                        st.error("Impossible d'envoyer : l'adresse e-mail est manquante pour cet enseignant. Importez un fichier de mapping avec la colonne 'Email'.")
                    elif not sender_email or not sender_password:
                        st.error("Les paramètres SMTP ne sont pas configurés.")
                    else:
                        try:
                            msg = MIMEMultipart()
                            msg['From'] = sender_email
                            msg['To'] = prof_email
                            msg['Subject'] = "Plateforme de gestion des EDTs-S2-2026 - Votre planning de surveillance"

                            corps_tableau = df_prof[['Enseignements', 'Horaire', 'Jours', 'Lieu', 'Promotion', 'Groupe']].to_html(index=False)
                            html_content = f"<p>{custom_msg.replace(chr(10), '<br>')}</p>{corps_tableau}"
                            msg.attach(MIMEText(html_content, 'html'))

                            server = smtplib.SMTP(smtp_server, smtp_port)
                            server.starttls()
                            server.login(sender_email, sender_password)
                            server.sendmail(sender_email, prof_email, msg.as_string())
                            server.quit()

                            st.success(f"E-mail envoyé avec succès à {selected_prof} ({prof_email}) !")
                        except Exception as e:
                            st.error(f"Erreur lors de l'envoi : {e}")

            # --- 2. ENVOI PAR GROUPE ---
            with tab_groupe:
                st.subheader("Envoi Groupé à tous les surveillants")
                st.info("Cette action va envoyer un e-mail à chaque surveillant listé dans les attributions, avec son planning personnel.")

                if st.button("Lancer l'envoi groupé", key="email_btn_group"):
                    if not sender_email or not sender_password:
                        st.error("Les paramètres SMTP ne sont pas configurés.")
                    else:
                        try:
                            server = smtplib.SMTP(smtp_server, smtp_port)
                            server.starttls()
                            server.login(sender_email, sender_password)

                            barre_progression = st.progress(0)
                            total = len(enseignants_list)
                            envoyes = 0
                            echecs = 0

                            for i, prof in enumerate(enseignants_list):
                                df_prof = df[df['Enseignants'] == prof]
                                df_prof = df_prof.drop_duplicates(subset=['Enseignements', 'Horaire', 'Jours', 'Lieu', 'Promotion', 'Groupe'])

                                prof_email = df_prof['Email'].iloc[0] if not df_prof.empty and 'Email' in df_prof.columns and pd.notna(df_prof['Email'].iloc[0]) and str(df_prof['Email'].iloc[0]).strip() != '' else ""

                                if prof_email:
                                    msg = MIMEMultipart()
                                    msg['From'] = sender_email
                                    msg['To'] = prof_email
                                    msg['Subject'] = "Plateforme de gestion des EDTs-S2-2026 - Votre planning de surveillance"

                                    corps_tableau = df_prof[['Enseignements', 'Horaire', 'Jours', 'Lieu', 'Promotion', 'Groupe']].to_html(index=False)
                                    html_content = f"<p>Bonjour Pr./Dr. {prof},<br><br>Veuillez trouver ci-dessous votre planning de surveillance extrait de la Plateforme de gestion des EDTs-S2-2026-Département d'Électrotechnique-Faculté de génie électrique-UDL-SBA :</p>{corps_tableau}"
                                    msg.attach(MIMEText(html_content, 'html'))

                                    server.sendmail(sender_email, prof_email, msg.as_string())
                                    envoyes += 1
                                else:
                                    echecs += 1

                                barre_progression.progress((i + 1) / total)

                            server.quit()
                            st.success(f"✅ Envoi groupé terminé : {envoyes} e-mail(s) envoyé(s), {echecs} sans adresse e-mail.")
                        except Exception as e:
                            st.error(f"Erreur lors de l'envoi groupé : {e}")
if __name__ == "__main__":
    main()
